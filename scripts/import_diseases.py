import os
import django
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'HCManagementSystem.settings')
django.setup()

from core.models import Disease

def import_diseases_from_excel(filename):
    excel_path = os.path.join(BASE_DIR, 'static', filename)
    print(f"Usando archivo: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[3]]
    print(f"Cabeceras: {headers}")
    idx_cod3 = headers.index('COD_3')
    idx_cod4 = headers.index('COD_4')
    idx_desc4 = headers.index('DESCRIPCION CODIGOS DE CUATRO CARACTERES')

    last_code_3 = None
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        code_3 = row[idx_cod3]
        code_4 = row[idx_cod4]
        name = row[idx_desc4]

        if not code_3:
            code_3 = last_code_3
        else:
            last_code_3 = code_3

        if code_3 and len(str(code_3)) > 4:
            print(f"Fila {i}: code_3 demasiado largo ({code_3}), saltando...")
            continue

        if not (code_3 and code_4 and name):
            print(f"Fila {i} incompleta, saltando...")
            continue
        Disease.objects.update_or_create(
            code_4=code_4,
            defaults={
                'code_3': code_3,
                'name': name
            }
        )
        print(f"Importado: {code_4} - {name}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python -m core.import_diseases CIE10.xlsx")
    else:
        import_diseases_from_excel(sys.argv[1])